
import pandas as pd
import openpyxl

file_path = '/Users/emre.yalciner/Desktop/Aktif Cimnastik Sistemleri/sinav/Sınav Sonuç Değerlendirme exc new (2).xlsx'

try:
    # Load workbook using openpyxl to check for formulas
    wb = openpyxl.load_workbook(file_path, data_only=False)
    print("Sheet Names:", wb.sheetnames)

    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        
        # Print first 10 rows to see structure and formulas
        print("First 10 rows (values & formulas):")
        for i, row in enumerate(ws.iter_rows(max_row=10)):
            row_data = []
            for cell in row:
                val = cell.value
                # If it looks like a formula, print it
                if isinstance(val, str) and val.startswith('='):
                    row_data.append(f"FORMULA: {val}")
                else:
                    row_data.append(str(val))
            print(f"Row {i+1}:", row_data)
            
    # Load data with pandas to see actual values easily
    print("\n--- Data Preview with Pandas (First 5 rows of each sheet) ---")
    xls = pd.ExcelFile(file_path)
    for sheet_name in xls.sheet_names:
        print(f"\nSheet: {sheet_name}")
        df = pd.read_excel(xls, sheet_name=sheet_name)
        print(df.head().to_string())
        print("-" * 20)

except Exception as e:
    print(f"Error analyzing Excel file: {e}")
