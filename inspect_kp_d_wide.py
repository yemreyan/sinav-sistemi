import openpyxl

file_path = 'Sınav Sonuç Değerlendirme exc new (2).xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

target_sheet = 'KP_D'
if target_sheet in wb.sheetnames:
    print(f"\n--- Inspecting {target_sheet} (Columns D-Z) ---")
    sheet = wb[target_sheet]
    
    # Check headers first (Rows 1-5)
    print("--- Headers ---")
    for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=30):
        vals = [str(c.value) if c.value is not None else "" for c in row]
        print(" | ".join(vals))

    # Check formulas in data rows (Row 6)
    print("\n--- Data Row 6 Formulas ---")
    for row in sheet.iter_rows(min_row=6, max_row=6, min_col=1, max_col=30):
        vals = [str(c.value) if c.value is not None else "" for c in row]
        print(" | ".join(vals))

else:
    print("Sheet KP_D not found.")
