
import pandas as pd
import openpyxl
import json

file_path = '/Users/emre.yalciner/Desktop/Aktif Cimnastik Sistemleri/sinav/Sınav Sonuç Değerlendirme exc new (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Tablo_E']
    
    # Row 1 is headers (Deviations): 0, 0.1, ...
    # Col A is Row Indexes (Expert Scores): 0, 0.1, ...
    
    matrix = {}
    
    # Get Headers (Deviations) - skipping A1 (None)
    headers = []
    for cell in ws[1]:
        if cell.column == 1: continue # Skip A1
        # Convert to float to avoid floating point issues (e.g. 0.300000004)
        val = cell.value
        if val is not None:
             headers.append(round(float(val), 1))
             
    # Iterate rows
    # Min row 2 because row 1 is header
    for row in ws.iter_rows(min_row=2):
        row_label = row[0].value
        if row_label is None: continue
        
        try:
            r_key = str(round(float(row_label), 1))
        except:
            continue
            
        row_data = {}
        for i, cell in enumerate(row[1:]): # Skip first col
            if i < len(headers):
                dev_key = str(headers[i])
                val = cell.value
                if val is not None:
                    row_data[dev_key] = val
                else:
                    row_data[dev_key] = 0 # Default to 0 if empty
        
        matrix[r_key] = row_data

    # Check the user's specific case: Expert 2.8, Dev 0.6
    print("User Case Check (Exp 2.8, Dev 0.6):", matrix.get("2.8", {}).get("0.6", "Not Found"))
    
    # Save to file
    with open('scoring_matrix.json', 'w') as f:
        json.dump(matrix, f)
        
    print("Matrix saved to scoring_matrix.json")

except Exception as e:
    print("Error:", e)
