import openpyxl

file_path = 'Sınav Sonuç Değerlendirme exc new (2).xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

print("All Sheets:", wb.sheetnames)

target_sheet = 'KP_D'
if target_sheet not in wb.sheetnames:
    print(f"Sheet {target_sheet} not found! Checking similar names...")
    for s in wb.sheetnames:
        if 'KP' in s or 'D' in s:
            print(f"Potential match: {s}")
    # Fallback to first sheet or specific logic if needed
    if 'ATM_D' in wb.sheetnames: target_sheet = 'ATM_D'
    elif 'DENGE_D' in wb.sheetnames: target_sheet = 'DENGE_D'
    elif 'YER_D' in wb.sheetnames: target_sheet = 'YER_D'

if target_sheet in wb.sheetnames:
    print(f"\n--- Inspecting {target_sheet} ---")
    sheet = wb[target_sheet]
    for row in sheet.iter_rows(min_row=1, max_row=20, max_col=15):
        vals = []
        for c in row:
            val = c.value
            # formatting for easier reading
            if val is None: vals.append("")
            else: vals.append(str(val))
        print(" | ".join(vals))
else:
    print("No relevant D-score sheet found.")
