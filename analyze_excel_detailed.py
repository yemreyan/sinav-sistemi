
import pandas as pd
import openpyxl

file_path = '/Users/emre.yalciner/Desktop/Aktif Cimnastik Sistemleri/sinav/Sınav Sonuç Değerlendirme exc new (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("ALL Sheet Names:", wb.sheetnames)

    if 'Tablo_E' in wb.sheetnames:
        print("\n--- Tablo_E Structure ---")
        ws = wb['Tablo_E']
        # Print first row (headers)
        headers = [c.value for c in ws[1]]
        print("Headers (Row 1):", headers[:20], "...")
        
        # Print first column (Row labels)
        row_labels = [r[0].value for r in ws.iter_rows(min_row=2, max_col=1)]
        print("Row Labels (Col A):", row_labels[:20], "...")

        # Print a sample matrix chunk
        print("Sample Data (Rows 2-7, Cols A-F):")
        for row in ws.iter_rows(min_row=2, max_row=7, max_col=6):
            print([c.value for c in row])

    # Check for D-related tables
    d_tables = [s for s in wb.sheetnames if 'D' in s and 'Tablo' in s]
    if d_tables:
        print(f"\nFound D Tables: {d_tables}")
        
    if 'Uzman_D' in wb.sheetnames:
        print("\n--- Uzman_D Preview ---")
        ws = wb['Uzman_D']
        for row in ws.iter_rows(max_row=5):
             print([c.value for c in row])

except Exception as e:
    print(f"Error: {e}")
