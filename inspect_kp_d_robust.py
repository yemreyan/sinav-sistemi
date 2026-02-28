import openpyxl
import glob

print("Searching for xlsx files...")
files = glob.glob('*.xlsx')
target_file = None

for f in files:
    if f.startswith("~$"): continue
    if "Sınav" in f or "Sonuc" in f or "Soınav" in f or "new" in f:
         target_file = f
         break

if target_file:
    print(f"Opening file: {target_file}")
    wb = openpyxl.load_workbook(target_file, data_only=False)

    target_sheet = 'KP_D'
    if target_sheet in wb.sheetnames:
        print(f"\n--- Inspecting {target_sheet} Columns 16-30 ---")
        sheet = wb[target_sheet]
        
        # Rows 1 to 10, Cols 16 to 30
        for r in range(1, 11):
            vals = []
            for c in range(16, 31):
                val = sheet.cell(row=r, column=c).value
                vals.append(str(val) if val is not None else "")
            print(f"Row {r}: " + " | ".join(vals))
    else:
        print(f"Sheet {target_sheet} not found in {wb.sheetnames}")
else:
    print("No suitable file found.")
